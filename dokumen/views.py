from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import FileResponse, Http404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Dokumen, Laporan
from .forms import DokumenForm, LaporanForm
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError
from django.utils.timezone import now
import json

@login_required(login_url='login')
def unggah_laporan(request, dokumen_id):
    dokumen = get_object_or_404(Dokumen, id=dokumen_id)
    laporan, created = Laporan.objects.get_or_create(dokumen=dokumen)

    if request.method == "POST":
        form = LaporanForm(request.POST, request.FILES, instance=laporan, dokumen=dokumen)
        if form.is_valid():
            if 'file' in request.FILES:
                form.save()
                dokumen.laporan_diunggah = True
                dokumen.status = "Sudah Diunggah"
                dokumen.save()
                messages.success(request, "Laporan berhasil diperbarui!")
                return redirect("daftar_dokumen")
            else:
                messages.error(request, "Anda harus mengunggah file laporan terlebih dahulu.")
        else:
            messages.error(request, "Terjadi kesalahan. Periksa kembali data yang Anda masukkan.")
    else:
        form = LaporanForm(instance=laporan, dokumen=dokumen)

    return render(request, "dokumen/unggah_laporan.html", {"form": form, "dokumen": dokumen})


@login_required(login_url='login')
def unggah_dokumen(request):
    if request.method == "POST":
        form = DokumenForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                dokumen = form.save(commit=False) 
                dokumen.user = request.user  
                dokumen.save()

                tim_audit_data = request.POST.get("tim_audit")
                if tim_audit_data:
                    try:
                        dokumen.tim_audit = json.loads(tim_audit_data)
                    except json.JSONDecodeError:
                        dokumen.tim_audit = []
                    dokumen.save()
                
                messages.success(request, "Surat Tugas berhasil diunggah.")
                return redirect("daftar_dokumen")
            except IntegrityError:
                messages.error(request, "Nomor surat sudah digunakan. Gunakan nomor yang berbeda.")
        else:
            print("Form Errors:", form.errors)
            messages.error(request, "Terjadi kesalahan, pastikan semua data telah diisi dengan benar.")
    else:
        form = DokumenForm()

    return render(request, "dokumen/unggah_dokumen.html", {"form": form,})


@login_required
def detail_dokumen(request, dokumen_id):
    dokumen = get_object_or_404(Dokumen, pk=dokumen_id)
    laporan = Laporan.objects.filter(dokumen=dokumen).first()

    print("Tim Audit yang dikirim ke template:", dokumen.tim_audit)

    next_page = request.GET.get("next", request.META.get("HTTP_REFERER", "/"))

    return render(request, "dokumen/detail_dokumen.html", {
        "dokumen": dokumen,
        "laporan": laporan,
        "next_page": next_page,
    })

@login_required
def unduh_surat_tugas(request, dokumen_id):
    
    dokumen = get_object_or_404(Dokumen, id=dokumen_id)

    if not dokumen.file or not dokumen.file.name:
        messages.error(request, "File Surat Tugas tidak tersedia.")
        return redirect("daftar_dokumen")

    return FileResponse(dokumen.file.open("rb"), as_attachment=True)

@login_required
def unduh_laporan(request, laporan_id):
    
    laporan = get_object_or_404(Laporan, id=laporan_id)

    if not laporan.file or not laporan.file.name:
        messages.error(request, "File Laporan tidak tersedia.")
        return redirect("daftar_dokumen")

    return FileResponse(laporan.file.open("rb"), as_attachment=True)


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from .models import Dokumen 

def login_view(request):
    if request.method == "POST":
        username = request.POST['username']
        password = request.POST['password']
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            if user.is_superuser or user.is_staff:
                return redirect('admin_dashboard')
            else:
                return redirect('dashboard')
        else:
            messages.error(request, "Username atau password salah!")

    return render(request, 'login.html')

def logout_view(request):
    logout(request)
    return redirect('login')

@login_required(login_url='login')
def profil(request):
    user = request.user
    return render(request, 'profil.html', {'user': user})


def register_view(request):
    if request.method == "POST":
        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        confirm_password = request.POST['confirm_password']

        if password == confirm_password:
            if User.objects.filter(username=username).exists():
                messages.error(request, "Username sudah digunakan.")
            elif User.objects.filter(email=email).exists():
                messages.error(request, "Email sudah digunakan.")
            else:
                user = User.objects.create_user(username=username, email=email, password=password)
                messages.success(request, "Pendaftaran berhasil! Silakan login.")
                return redirect('login')
        else:
            messages.error(request, "Password tidak cocok.")

    return render(request, 'register.html')


def is_admin(user):
    return user.is_superuser or user.is_staff

@user_passes_test(is_admin)
def hapus_pengguna(request, user_id):
    user = get_object_or_404(User, id=user_id)
    user.delete()
    messages.success(request, "Pengguna berhasil dihapus.")
    return redirect('admin_dashboard')

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def create_user(request):
    if request.method == "POST":
        full_name = request.POST['full_name']
        first_name, last_name = full_name.split(' ', 1) if ' ' in full_name else (full_name, '')

        username = request.POST['username']
        email = request.POST['email']
        password = request.POST['password']
        is_staff = request.POST.get('is_staff', False) == "on" 

        if User.objects.filter(username=username).exists():
            messages.error(request, "Username sudah digunakan.")
        elif User.objects.filter(email=email).exists():
            messages.error(request, "Email sudah digunakan.")
        else:
            user = User.objects.create_user(username=username, email=email, password=password)
            user.first_name = first_name
            user.last_name = last_name
            user.is_staff = is_staff
            user.save()
            messages.success(request, "Akun berhasil dibuat!")
            return redirect('admin_dashboard')

    return render(request, 'dokumen/create_user.html')

@login_required(login_url='login')
@user_passes_test(is_admin, login_url='dashboard')
def update_user(request, user_id):
    user = User.objects.get(id=user_id)

    if request.method == "POST":
        full_name = request.POST['full_name']
        first_name, last_name = full_name.split(' ', 1) if ' ' in full_name else (full_name, '')

        user.username = request.POST['username']
        user.email = request.POST['email']
        user.first_name = first_name
        user.last_name = last_name
        user.is_staff = request.POST.get('is_staff', False) == "on"

        if 'password' in request.POST and request.POST['password']:
            user.set_password(request.POST['password'])

        user.save()
        messages.success(request, "Akun berhasil diperbarui!")
        return redirect('admin_dashboard')

    return render(request, 'dokumen/update_user.html', {'user': user})


@user_passes_test(is_admin)
def daftar_dokumen_admin(request):

    nomor_surat_query = request.GET.get("nomor_surat", "")
    tanggal_surat_query = request.GET.get("tanggal_surat", "")
    irban_query = request.GET.get("irban", "")

    # Filter pencarian
    dokumen_list = Dokumen.objects.select_related("user", "laporan").all()
    
    if nomor_surat_query:
        dokumen_list = dokumen_list.filter(nomor_surat__icontains=nomor_surat_query)
    if tanggal_surat_query:
        dokumen_list = dokumen_list.filter(tanggal_surat=tanggal_surat_query)
    if irban_query:
        dokumen_list = dokumen_list.filter(irban__icontains=irban_query)

    dokumen_list = dokumen_list.order_by("-tanggal_surat", "-id")


    paginator = Paginator(dokumen_list, 10)
    page_number = request.GET.get("page")

    try:
        dokumen_page = paginator.page(page_number)
    except (EmptyPage, PageNotAnInteger):
        dokumen_page = paginator.page(1)

    return render(request, 'dokumen/daftar_dokumen_admin.html', {'dokumen_list': dokumen_list})

@user_passes_test(is_admin)
def hapus_dokumen(request, dokumen_id):
    dokumen = get_object_or_404(Dokumen, id=dokumen_id)
    dokumen.delete()
    messages.success(request, "Dokumen berhasil dihapus.")
    return redirect('admin_daftar_dokumen')

from django.shortcuts import render
from .models import Dokumen

def daftar_dokumen(request):
    today = now().date()
    nomor_surat_query = request.GET.get("nomor_surat", "")
    tahun_choices = Dokumen.objects.values_list('tanggal_surat__year', flat=True).distinct()
    irban_choices = Dokumen.objects.values_list('irban', flat=True).distinct()

    tahun = request.GET.get('tahun')
    irban = request.GET.get('irban')

    dokumen_list = Dokumen.objects.filter(user=request.user)

    # Filter berdasarkan tahun dan irban jika ada
    if nomor_surat_query:
        dokumen_list = dokumen_list.filter(nomor_surat__icontains=nomor_surat_query)
    if tahun:
        dokumen_list = dokumen_list.filter(tanggal_surat__year=tahun)
    if irban:
        dokumen_list = dokumen_list.filter(irban=irban)

    dokumen_list = dokumen_list.order_by("-tanggal_surat", "-id")

    # Periksa output SQL yang dihasilkan
    print(dokumen_list.query)

    context = {
        'dokumen_list': dokumen_list,
        'tahun_choices': tahun_choices,
        'irban_choices': irban_choices,
        'today': today,
    }
    
    return render(request, 'dokumen/daftar_dokumen.html', context)

import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment
from django.http import HttpResponse
from .models import Dokumen

def ekspor_excel(request):
    # Ambil data sesuai filter yang dipilih
    tahun = request.GET.get('tahun')
    irban = request.GET.get('irban')

    dokumen_list = Dokumen.objects.filter(user=request.user)
    if tahun:
        dokumen_list = dokumen_list.filter(tanggal_surat__year=tahun)
    if irban:
        dokumen_list = dokumen_list.filter(irban=irban)

    # Membuat workbook baru
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Dokumen"

    # Gaya Border (garis tabel)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))

    # Header kolom
    headers = [
        'Nomor Surat', 'Tanggal Surat', 'Irban', 'Tim Audit (Nama & Jabatan)', 
        'Uraian', 'Nomor Laporan', 'Tanggal Laporan', 'Tanggal Masuk Laporan'
    ]
    ws.append(headers)

    # Gaya header (Bold + Center + Border)
    for col_num, col_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    row_num = 2  # Mulai dari baris kedua (baris pertama adalah header)

    for dokumen in dokumen_list:
        laporan = dokumen.laporan if hasattr(dokumen, 'laporan') else None

        nomor_laporan = laporan.nomor_laporan if laporan else "-"
        tanggal_laporan = laporan.tanggal_laporan.strftime('%d-%m-%Y') if laporan and laporan.tanggal_laporan else "-"
        tanggal_masuk_surat = laporan.tanggal_masuk_surat.strftime('%d-%m-%Y') if laporan and laporan.tanggal_masuk_surat else "-"
        tanggal_surat = dokumen.tanggal_surat.strftime('%d-%m-%Y') if dokumen.tanggal_surat else "-"

        tim_audit_list = dokumen.tim_audit if dokumen.tim_audit else []
        
        # Jika ada anggota tim audit, gabungkan nama dan jabatan ke dalam satu sel (dengan format terurut ke bawah)
        if tim_audit_list:
            tim_audit_str = "\n".join([f"{item['nama']} - {item['jabatan']}" for item in tim_audit_list])
        else:
            tim_audit_str = "-"  # Jika tidak ada anggota tim audit

        row_data = [
            dokumen.nomor_surat,
            tanggal_surat,
            dokumen.irban,
            tim_audit_str,  # Format anggota tim audit ke bawah
            dokumen.uraian,
            nomor_laporan,
            tanggal_laporan,
            tanggal_masuk_surat,
        ]
        ws.append(row_data)

        # Terapkan border, format alignment (rata tengah), dan wrap text
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Khusus untuk kolom "Tim Audit", teks tetap wrap text dengan posisi di atas
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)

        row_num += 1  # Pindah ke baris berikutnya

    # *Menyesuaikan lebar kolom secara otomatis agar data panjang tidak terpotong*
    for col_num, col_name in enumerate(headers, 1):
        max_length = max(len(str(ws.cell(row=row, column=col_num).value)) for row in range(1, row_num + 1))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = max(15, min(max_length + 5, 50))

    # Menyusun response untuk mendownload file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=daftar_dokumen.xlsx'
    wb.save(response)

    return response
